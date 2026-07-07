import csv
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S").rstrip(" 00:00:00")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def main():
    if len(sys.argv) < 2:
        raise SystemExit("xlsx path is required")
    path = Path(sys.argv[1])
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.active
    writer = csv.writer(sys.stdout, lineterminator="\n")
    for row in worksheet.iter_rows(values_only=True):
        writer.writerow([cell_text(value) for value in row])


if __name__ == "__main__":
    main()
