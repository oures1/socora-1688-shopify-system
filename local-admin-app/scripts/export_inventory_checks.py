#!/usr/bin/env python3
import json
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def cell_value(value):
    if value is None:
        return ""
    return value


def set_header(row):
    fill = PatternFill("solid", fgColor="E5E7EB")
    for cell in row:
        cell.font = Font(bold=True, color="111827")
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def autosize(ws, max_width=64):
    for col in ws.columns:
        width = 10
        letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[letter].width = width


def status_fill(status):
    if status in ("available",):
        return PatternFill("solid", fgColor="DCFCE7")
    if status in ("partial",):
        return PatternFill("solid", fgColor="FEF3C7")
    if status in ("out", "error", "protected", "link_broken"):
        return PatternFill("solid", fgColor="FEE2E2")
    return PatternFill("solid", fgColor="F3F4F6")


def main():
    if len(sys.argv) != 3:
        print("Usage: export_inventory_checks.py input.json output.xlsx", file=sys.stderr)
        return 2

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    data = json.loads(input_path.read_text(encoding="utf-8"))
    targets = data.get("targets") or []
    checks = data.get("checks") or []
    summary = data.get("summary") or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "概要"
    ws.append(["項目", "件数"])
    set_header(ws[1])
    rows = [
        ("対象商品", summary.get("targetCount", 0)),
        ("確認済み", summary.get("checkedCount", 0)),
        ("在庫あり/一部在庫なし", summary.get("availableCount", 0)),
        ("在庫なし", summary.get("outCount", 0)),
        ("要確認/取得失敗", summary.get("errorCount", 0)),
    ]
    for row in rows:
        ws.append(list(row))
    autosize(ws)

    ws_products = wb.create_sheet("商品別在庫")
    product_headers = [
        "管理番号", "商品名", "在庫判定", "総在庫", "SKU取得数", "在庫なしSKU",
        "最終確認日時", "Shopify反映日時", "Shopify反映SKU", "Shopify未照合",
        "照合SKU", "更新SKU", "変更内容", "仕入れURL", "Shopify URL", "エラー",
    ]
    ws_products.append(product_headers)
    set_header(ws_products[1])
    for item in targets:
        latest = item.get("latestCheck") or item.get("lastInventoryCheck") or {}
        status = latest.get("status") or ""
        ws_products.append([
            item.get("productNo", ""),
            item.get("title", ""),
            latest.get("statusLabel") or status or "未確認",
            cell_value(latest.get("totalStock")),
            cell_value(latest.get("knownRows")),
            cell_value(latest.get("outRows")),
            latest.get("checkedAt", ""),
            latest.get("shopifyAppliedAt", ""),
            cell_value(latest.get("shopifyAppliedCount")),
            " / ".join(latest.get("shopifyMissing") or []),
            cell_value(latest.get("matchedVariants")),
            cell_value(latest.get("updatedVariants")),
            latest.get("changeSummary", ""),
            item.get("sourceUrl", ""),
            item.get("shopifyUrl", ""),
            latest.get("error", ""),
        ])
        for cell in ws_products[ws_products.max_row]:
            cell.fill = status_fill(status)
    autosize(ws_products)
    ws_products.freeze_panes = "A2"

    ws_sku = wb.create_sheet("SKU別在庫")
    sku_headers = [
        "確認日時", "管理番号", "商品名", "判定", "元カラー", "元サイズ", "規格",
        "在庫数", "在庫原文", "価格", "仕入れURL",
    ]
    ws_sku.append(sku_headers)
    set_header(ws_sku[1])
    for check in checks:
        rows = check.get("rows") or []
        if not rows:
            ws_sku.append([
                check.get("checkedAt", ""),
                check.get("productNo", ""),
                check.get("title", ""),
                check.get("statusLabel") or check.get("status", ""),
                "", "", "", "", "", "", check.get("sourceUrl", ""),
            ])
            continue
        for row in rows:
            ws_sku.append([
                check.get("checkedAt", ""),
                check.get("productNo", ""),
                check.get("title", ""),
                check.get("statusLabel") or check.get("status", ""),
                row.get("color", ""),
                row.get("size", ""),
                row.get("spec", ""),
                cell_value(row.get("stockNumber")),
                row.get("stockRaw", ""),
                row.get("price", ""),
                check.get("sourceUrl", ""),
            ])
    autosize(ws_sku)
    ws_sku.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
