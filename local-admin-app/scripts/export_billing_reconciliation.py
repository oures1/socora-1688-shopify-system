import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


HEADERS = [
    "ステータス",
    "カスタマー注文番号",
    "受取人",
    "BANRI注文番号",
    "物流番号",
    "請求書番号",
    "管理番号",
    "SKU",
    "商品名",
    "数量",
    "商品代金",
    "国内送料",
    "BANRI手数料",
    "国際送料按分",
    "その他費用",
    "原価合計",
    "売上",
    "行粗利",
    "行利益率",
    "1個原価",
    "1個原価率",
    "1個粗利",
    "1個利益率",
    "同梱件数",
    "差額/要確認",
    "お知らせ",
    "メモ",
]


def yen(value):
    try:
        return int(round(float(value or 0)))
    except (TypeError, ValueError):
        return 0


def main():
    if len(sys.argv) < 3:
        raise SystemExit("input json and output xlsx are required")
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    payload = json.loads(input_path.read_text(encoding="utf-8"))
    items = payload.get("items", [])
    summary = payload.get("summary", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "請求突合"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1, len(items) + 1)}"

    header_fill = PatternFill("solid", fgColor="D9E2F3")
    thin = Side(style="thin", color="B7C0CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_no, item in enumerate(items, 2):
        values = [
            item.get("status", ""),
            item.get("customerOrderNo", ""),
            item.get("recipientName", ""),
            item.get("banriOrderNo", ""),
            item.get("logisticsNo", ""),
            item.get("invoiceNumber", ""),
            item.get("productNo", ""),
            item.get("sku", ""),
            item.get("productName", ""),
            item.get("quantity", 0),
            yen(item.get("productCostJpy")),
            yen(item.get("domesticShippingJpy")),
            yen(item.get("workFeeJpy")),
            yen(item.get("allocatedInternationalShippingJpy")),
            yen(item.get("otherFeeJpy")),
            yen(item.get("totalCostJpy")),
            yen(item.get("salesJpy")),
            yen(item.get("grossProfitJpy")),
            item.get("grossMarginPct", 0),
            yen(item.get("unitCostJpy")),
            item.get("unitCostRatePct", 0),
            yen(item.get("unitGrossProfitJpy")),
            item.get("unitGrossMarginPct", 0),
            item.get("coShipmentCount", 0),
            " / ".join(item.get("issues", []) or []),
            " / ".join(item.get("notices", []) or []),
            item.get("note", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_no, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=col in (9, 25, 26, 27))
            if col in (11, 12, 13, 14, 15, 16, 17, 18, 20, 22):
                cell.number_format = '#,##0'
            if col in (19, 21, 23):
                cell.number_format = '0.0%'

    widths = [14, 18, 16, 18, 16, 18, 12, 18, 34, 8, 12, 12, 12, 14, 12, 12, 12, 12, 10, 12, 10, 12, 10, 10, 34, 28, 24]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    summary_ws = wb.create_sheet("サマリー")
    summary_rows = [
        ("明細数", summary.get("itemCount", 0)),
        ("要対応", summary.get("actionCount", 0)),
        ("未請求", summary.get("unbilledCount", 0)),
        ("追跡なし", summary.get("missingTrackingCount", 0)),
        ("低粗利", summary.get("lowMarginCount", 0)),
        ("品番のみ", summary.get("itemNoOnlyCount", 0)),
        ("確定候補", summary.get("finalCandidateCount", 0)),
        ("確認済み", summary.get("confirmedCount", 0)),
        ("売上", yen(summary.get("totalSalesJpy"))),
        ("原価", yen(summary.get("totalCostJpy"))),
        ("粗利", yen(summary.get("totalGrossProfitJpy"))),
    ]
    for row_no, (label, value) in enumerate(summary_rows, 1):
        summary_ws.cell(row=row_no, column=1, value=label).font = Font(bold=True)
        summary_ws.cell(row=row_no, column=2, value=value)
    summary_ws.column_dimensions["A"].width = 16
    summary_ws.column_dimensions["B"].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    main()
