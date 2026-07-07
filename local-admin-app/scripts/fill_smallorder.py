import copy
import base64
import io
import json
import sys
import tempfile
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage


def copy_cell_style(source, target):
    if source.has_style:
        target.font = copy.copy(source.font)
        target.fill = copy.copy(source.fill)
        target.border = copy.copy(source.border)
        target.alignment = copy.copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy.copy(source.protection)


def write_link(cell, value, label=None):
    text = str(value or "")
    cell.value = label or text
    if text.startswith(("http://", "https://")):
        cell.hyperlink = text
        cell.style = "Hyperlink"


def read_image_bytes(url):
    text = str(url or "").strip()
    if not text:
        return None
    if text.startswith("data:"):
        header, _, payload = text.partition(",")
        if "svg" in header.lower():
            return None
        if ";base64" in header.lower():
            return base64.b64decode(payload)
        return urllib.parse.unquote_to_bytes(payload)
    if text.startswith(("http://", "https://")):
        req = urllib.request.Request(
            text,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            },
        )
        with urllib.request.urlopen(req, timeout=12) as response:
            content_type = response.headers.get("Content-Type", "").lower()
            if "svg" in content_type:
                return None
            return response.read(8_000_000)
    return None


def prepare_image_file(url, temp_dir):
    try:
        raw = read_image_bytes(url)
        if not raw:
            return None
        with PILImage.open(io.BytesIO(raw)) as source:
            image = source.convert("RGBA")
            image.thumbnail((64, 64))
            canvas = PILImage.new("RGBA", (64, 64), (255, 255, 255, 0))
            left = (64 - image.width) // 2
            top = (64 - image.height) // 2
            canvas.alpha_composite(image, (left, top))
            output = Path(temp_dir) / f"image_{len(list(Path(temp_dir).iterdir()))}.png"
            canvas.save(output, "PNG")
            return output
    except Exception:
        return None


def column_width_to_pixels(width):
    if width is None:
        width = 8.43
    width = float(width)
    if width <= 0:
        return 0
    if width < 1:
        return int(width * 12)
    return int(width * 7 + 5)


def row_height_to_pixels(height):
    if height is None:
        height = 15
    return int(float(height) * 96 / 72)


def fit_image_to_cell(ws, image, cell_ref, margin_px=8, offset_right_px=4, offset_down_px=3):
    row_no, col_no = coordinate_to_tuple(cell_ref)
    col_letter = get_column_letter(col_no)
    cell_width = column_width_to_pixels(ws.column_dimensions[col_letter].width)
    cell_height = row_height_to_pixels(ws.row_dimensions[row_no].height)
    max_width = max(1, cell_width - margin_px * 2)
    max_height = max(1, cell_height - margin_px * 2)

    ratio = min(max_width / image.width, max_height / image.height, 1)
    image.width = max(1, int(image.width * ratio))
    image.height = max(1, int(image.height * ratio))

    # BANRI側のアップロード判定で、写真がセル境界に近すぎると失敗することがある。
    # 中央寄せを維持しつつ、少しだけ右下へ逃がしてセル内に確実に収める。
    left = max(0, min(cell_width - image.width, int((cell_width - image.width) / 2) + offset_right_px))
    top = max(0, min(cell_height - image.height, int((cell_height - image.height) / 2) + offset_down_px))
    image.anchor = OneCellAnchor(
        _from=AnchorMarker(
            col=col_no - 1,
            colOff=pixels_to_EMU(left),
            row=row_no - 1,
            rowOff=pixels_to_EMU(top),
        ),
        ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(image.width),
            cy=pixels_to_EMU(image.height),
        ),
    )


def add_image(ws, cell_ref, image_path):
    if not image_path:
        return
    image = ExcelImage(str(image_path))
    fit_image_to_cell(ws, image, cell_ref)
    ws.add_image(image)


def create_default_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "模板"
    headers = [
        "中国追跡番号",
        "写真",
        "* SKU/URL",
        "仕様/サイズ",
        "色",
        "* 数量",
        "単価(CNY)",
        "合計(CNY)",
        "* 配送先氏名",
        "* 電話番号",
        "* 配送先住所",
        "* 郵便番号",
        "メールアドレス",
        "商品番号",
        "備考",
        "御社管理用注文番号",
    ]
    widths = [16, 11, 32, 14, 12, 10, 13, 14, 16, 16, 25, 14, 24, 12, 12, 22]
    header_fill = PatternFill("solid", fgColor="B8C7E6")
    required_font = Font(bold=True, color="FF0000")
    normal_font = Font(bold=True, color="000000")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    for index, title in enumerate(headers, start=1):
        cell = ws.cell(1, index)
        cell.value = title
        cell.fill = header_fill
        cell.font = required_font if title.startswith("*") else normal_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = widths[index - 1]
    ws.row_dimensions[1].height = 28
    for col in range(1, 17):
        cell = ws.cell(2, col)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return wb


def main():
    if len(sys.argv) != 4:
        raise SystemExit("usage: fill_smallorder.py rows.json template.xlsx output.xlsx")

    rows_path = Path(sys.argv[1])
    template_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])
    rows = json.loads(rows_path.read_text(encoding="utf-8")).get("items", [])

    wb = load_workbook(template_path) if template_path.exists() else create_default_workbook()
    ws = wb["模板"] if "模板" in wb.sheetnames else wb[wb.sheetnames[0]]
    template_row = 2
    start_row = 3

    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)
    ws._images = []
    for col in range(1, 17):
        cell = ws.cell(template_row, col)
        cell.value = ""
        cell.hyperlink = None
        cell.comment = None
    ws.row_dimensions[template_row].height = 54

    with tempfile.TemporaryDirectory() as temp_dir:
        for offset, item in enumerate(rows):
            row_no = start_row + offset
            if row_no > ws.max_row:
                ws.insert_rows(row_no)
            for col in range(1, 17):
                copy_cell_style(ws.cell(template_row, col), ws.cell(row_no, col))
            ws.row_dimensions[row_no].height = 54

            qty = int(float(item.get("quantity") or 1))
            unit = float(item.get("unitCny") or 0)

            for col in range(1, 17):
                cell = ws.cell(row_no, col)
                cell.value = ""
                cell.hyperlink = None
                cell.comment = None

            ws.cell(row_no, 1).value = ""
            ws.cell(row_no, 2).value = ""
            add_image(ws, f"B{row_no}", prepare_image_file(item.get("imageUrl"), temp_dir))
            write_link(ws.cell(row_no, 3), item.get("sourceUrl"))
            ws.cell(row_no, 4).value = item.get("originalSize") or ""
            ws.cell(row_no, 5).value = item.get("originalColor") or ""
            ws.cell(row_no, 6).value = qty
            ws.cell(row_no, 7).value = unit
            ws.cell(row_no, 8).value = qty * unit
            ws.cell(row_no, 9).value = item.get("shippingName") or ""
            ws.cell(row_no, 10).value = item.get("shippingPhone") or ""
            ws.cell(row_no, 11).value = item.get("shippingAddress") or ""
            ws.cell(row_no, 12).value = item.get("shippingZip") or ""
            ws.cell(row_no, 13).value = item.get("email") or ""
            ws.cell(row_no, 14).value = item.get("productNo") or ""
            ws.cell(row_no, 15).value = ""
            order_name = str(item.get("orderName") or "").strip()
            ws.cell(row_no, 16).value = f"#{order_name}" if order_name and not order_name.startswith("#") else order_name

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)


if __name__ == "__main__":
    main()
